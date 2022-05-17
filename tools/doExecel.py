# -*- coding: utf-8 -*- 
# @Time : 2022/4/27 21:54 
# @Author : wxy
# @File : doExecel.py

from openpyxl import load_workbook,worksheet
from tools.read_config import ReadConfig
from tools import project_path
from tools.get_data import GetData

class DoExcel:
    @classmethod
    def get_data(cls,file_name):
        wb = load_workbook(file_name)
        test_data = []
        tel=getattr(GetData,'NoRegTel')  #从GetData拿到数据
        # print(type(tel))
        mode = eval(ReadConfig.get_config(project_path.case_config_path, "MODE", "mode"))
        for key in mode:  # 遍历这个存在配置文件里面的字典
            sheet = wb[key]  # key就是表单名
            if mode[key]=="all":
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['title'] = sheet.cell(i, 2).value
                    row_data['method'] = sheet.cell(i, 3).value
                    row_data['url'] = sheet.cell(i, 4).value
                    if sheet.cell(i, 5).value.find('${tel}')!=-1:
                        row_data['data']=sheet.cell(i, 5).value.replace('${tel}',str(tel))
                        tel+=1
                    elif sheet.cell(i, 5).value.find('${admin_tel}')!=-1:
                        row_data['data'] = sheet.cell(i, 5).value.replace('${admin_tel}', str(getattr(GetData,'admin_tel')))
                    elif sheet.cell(i, 5).value.find('${loan_member_id}') != -1:
                        row_data['data'] = sheet.cell(i, 5).value.replace('${loan_member_id}',str(getattr(GetData,'loan_member_id')))
                    elif sheet.cell(i, 5).value.find('${normal_tel}')!=-1:
                        row_data['data'] = sheet.cell(i, 5).value.replace('${normal_tel}', str(getattr(GetData,'normal_tel')))
                    else:
                        row_data['data'] = sheet.cell(i, 5).value
                    row_data['expect'] = sheet.cell(i, 6).value
                    row_data['sheet_name']=key
                    test_data.append(row_data)
                    cls.update_tel(tel+1,file_name,"teldata")
            else:
                for case_id in mode[key]:  #不等于all，就是一个空列表
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['title'] = sheet.cell(case_id+1, 2).value
                    row_data['method'] = sheet.cell(case_id+1, 3).value
                    row_data['url'] = sheet.cell(case_id+1, 4).value
                    if sheet.cell(case_id+1, 5).value.find('${tel}')!=-1:
                        row_data['data']=sheet.cell(case_id+1, 5).value.replace('${tel_1}',str(tel))
                        tel += 1
                    elif sheet.cell(case_id+1, 5).value.find('${admin_tel}') != -1:
                        row_data['data'] = sheet.cell(case_id+1, 5).value.replace('${admin_tel}',str(getattr(GetData, 'admin_tel')))
                    elif sheet.cell(case_id+1, 5).value.find('${loan_member_id}') != -1:
                        row_data['data'] = sheet.cell(case_id+1, 5).value.replace('${loan_member_id}',str(getattr(GetData, 'loan_member_id')))
                    elif sheet.cell(case_id+1, 5).value.find('${normal_tel}') != -1:
                        row_data['data'] = sheet.cell(case_id+1, 5).value.replace('${normal_tel}',str(getattr(GetData, 'normal_tel')))
                    else:
                        row_data['data'] = sheet.cell(case_id+1, 5).value
                    row_data['expect'] = sheet.cell(case_id+1, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    cls.update_tel(tel + 1, file_name, "teldata")
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):  # 专门写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存结果

    @classmethod
    def update_tel(cls,tel,file_name,sheet_name):
        wb=load_workbook(file_name)
        sheet= wb[sheet_name]
        sheet.cell(2,1).value=tel
        wb.save(file_name)

if __name__ == '__main__':
    res = DoExcel.get_data(project_path.test_case_path)
    print(res)
